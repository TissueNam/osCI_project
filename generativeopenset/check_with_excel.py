#!/usr/bin/env python
import argparse
import os
import sys
from pprint import pprint

import torch
import numpy as np
from torch.autograd import Variable
import torch.nn.functional as F
from PIL import Image

# For excel control
# import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Color
from openpyxl.drawing.image import Image
from PIL import Image as PIL_Image
from evaluation import plot_confusion_matrix

# DEFECT_CLASSES = ['Array', 'Gap', 'Particle', 'Scratch', 'Undefined']
save_path = './excel_check_fold'
failed_img_path = save_path + '/failed_img_fold'
img_name = "/confusion_matrix.png"

def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print ('Error: Creating directory. ' +  directory)
 
createFolder(save_path)
createFolder(failed_img_path)

def check_with_excel(networks, dataloader, **options):
    for net in networks.values():
        net.eval()

    netC = networks['classifier']    
    fold = options.get('fold', 'evaluation')

    classification_closed_correct = 0
    classification_total = 0

    predicted_array = torch.zeros(dataloader.dsf.count(dataloader.fold), dtype=int)
    class_labels_array = torch.zeros(dataloader.dsf.count(dataloader.fold))

    cnt = 0

    # Set excel environment
    DEFECT_CLASSES = dataloader.lab_conv.labels
    TRAINED_DEFECT_CLASSES = dataloader.lab_conv.labels + ['true label', 'predict']
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(TRAINED_DEFECT_CLASSES)
    sheet.title = "SoftMax Values"
    r_color = PatternFill(start_color='ff9999', end_color='ff9999', fill_type='solid')
    g_color = PatternFill(start_color='ccff99', end_color='ccff99', fill_type='solid')
    with torch.no_grad():
        for i, (images, class_labels) in enumerate(dataloader):
            images = images.unsqueeze(1)
            images = Variable(images, requires_grad=False).cuda()
            labels = Variable(class_labels, requires_grad=False).cuda()

            net_y = netC(images)
            class_predictions = F.softmax(net_y, dim=1)
            _, predicted = class_predictions.max(1)
            
            origin_data = net_y.cpu().numpy()
            softmax_data = class_predictions.cpu().numpy()

            for idx in range(softmax_data.shape[0]):
                
                processed_data = softmax_data[idx]
                processed_data = np.append(processed_data, [0, 0])
                processed_data = np.append(processed_data, origin_data[idx])

                # sheet.append(softmax_data[idx].tolist())
                sheet.append(processed_data.tolist())
                sheet.cell(row=sheet._current_row, column=(len(TRAINED_DEFECT_CLASSES)-1)).value = DEFECT_CLASSES[labels[idx]] # True
                sheet.cell(row=sheet._current_row, column=len(TRAINED_DEFECT_CLASSES)).value = DEFECT_CLASSES[predicted[idx]] # predict

                # sheet.cell(row=sheet._current_row, column=(len(TRAINED_DEFECT_CLASSES)-1)).value = DEFECT_CLASSES[labels[idx]] # True
                # color
                sheet.cell(row=sheet._current_row, column=len(TRAINED_DEFECT_CLASSES)).fill = g_color
                sheet.cell(row=sheet._current_row, column=(len(TRAINED_DEFECT_CLASSES)-1)).fill = g_color
                if (labels[idx] != predicted[idx]):
                    sheet.cell(row=sheet._current_row, column=len(TRAINED_DEFECT_CLASSES)).fill = r_color
                    sheet.cell(row=sheet._current_row, column=(len(TRAINED_DEFECT_CLASSES)-1)).fill = r_color
            
                    # Add Image
                    failed_img = images[idx].cpu().numpy()
                    failed_img = (failed_img-failed_img.min())/(failed_img.max()-failed_img.min())
                    failed_img = failed_img*255
                    failed_img = failed_img.reshape((512, 512)).astype(np.uint8)
                    file_name = failed_img_path + "/true_" + DEFECT_CLASSES[labels[idx]] + "_to_" + DEFECT_CLASSES[predicted[idx]] + "_" + str(cnt+idx+2) +".png"  # cnt: excel is 1-based and row(1) is classes'name
                    PIL_Image.fromarray(failed_img).save(file_name)
                    failed_img = Image(file_name)
                    failed_img.width = 150
                    failed_img.height = 150
                    save_image_cell = 'L' + str(sheet._current_row)
                    sheet.add_image(failed_img, save_image_cell)
               

            classification_closed_correct += sum(predicted.data == labels)
            classification_total += len(labels)

            # confusion map
            predicted_array[cnt:cnt+len(labels)] = predicted
            class_labels_array[cnt:cnt+len(labels)] = labels
            cnt += len(labels)
    print(cnt)
    basic_stat = {
        fold: {
            'closed_set_image_class_accuracy': float(classification_closed_correct) / (classification_total),
        }
    }

    # Add Confusion Map Image
    img_path = save_path + img_name
    plot_confusion_matrix(class_labels_array, predicted_array, pngName=img_path ,classes=dataloader.lab_conv.labels, title='Confusion matrix')
    img = Image(img_path)
    sheet.add_image(img, 'N4')

    file_name = save_path +'/check_softmax_val.xlsx'
    wb.save(file_name)

    return basic_stat

parser = argparse.ArgumentParser()
parser.add_argument('--result_dir', help='Output directory for images and model checkpoints [default: .]', default='.')
parser.add_argument('--epochs', type=int, default=1000, help='number of epochs to train for [default: 10]')
options = vars(parser.parse_args())

sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from dataloader import CustomDataloader, FlexibleCustomDataloader
from networks import build_VGG_basic_network, get_optimizers_basic
from options import load_options, get_current_epoch
from evaluation import draw_confidence_map, draw_basic_confidence_map

options = load_options(options)
dataloader = FlexibleCustomDataloader(fold='train', **options)
eval_dataloader = CustomDataloader(last_batch=True, shuffle=False, fold='test', **options)

networks = build_VGG_basic_network(dataloader.num_classes, **options)
optimizers = get_optimizers_basic(networks, finetune=False, **options)

# start_epoch = get_current_epoch(options['result_dir']) + 1

evalate = check_with_excel(networks, eval_dataloader, **options)

print(evalate)